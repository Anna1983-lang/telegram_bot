import logging
import os
import shutil
from datetime import datetime

from aiogram import Bot, Dispatcher, Router, F
from aiogram.filters import CommandStart, Command
from aiogram.types import Message, CallbackQuery, FSInputFile
from aiogram.fsm.storage.memory import MemoryStorage
from aiogram.fsm.state import State, StatesGroup
from aiogram.fsm.context import FSMContext

import openpyxl
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# ─── КОНФИГУРАЦИЯ ────────────────────────────────────────
logging.basicConfig(level=logging.INFO)
TOKEN = "8475192387:AAESFlpUUqJzlqPTQkcAv1sDVeZJSFOQV0w"
ADMIN_ID = 1227847495

POLICY_PDF = "policy.pdf"
CONSENT_PDF = "consent2.pdf"
OFFER_PDF = "ПУБЛИЧНОЕ ПРЕДЛОЖЕНИЕ (ОФЕРТА).pdf"
EXCEL_FILE = "consents.xlsx"

pdfmetrics.registerFont(TTFont("DejaVu", "DejaVuSans.ttf"))
pdfmetrics.registerFont(TTFont("DejaVu-Bold", "DejaVuSans-Bold.ttf"))

router = Router()
storage = MemoryStorage()

# ─── СОСТОЯНИЯ ДЛЯ FSM ──────────────────────────────────
class ConsentForm(StatesGroup):
    waiting_fullname = State()
    waiting_inn = State()

# ─── СТАРТОВОЕ СООБЩЕНИЕ ────────────────────────────────
AGREEMENT_TEXT = """
🔒 Согласие на обработку персональных данных

Ознакомьтесь с документами (PDF) ниже. Для согласия укажите ФИО и ИНН — это обязательно.
"""

@router.message(CommandStart())
async def start_handler(m: Message, state: FSMContext):
    kb = [
        [
            {"text": "📄 Политика (PDF)", "callback_data": "policy_pdf"},
            {"text": "📝 Согласие (PDF)", "callback_data": "consent_pdf"},
            {"text": "📜 Оферта (PDF)", "callback_data": "offer_pdf"},
        ],
        [
            {"text": "✅ Согласен", "callback_data": "agree"},
            {"text": "❌ Не согласен", "callback_data": "disagree"}
        ]
    ]
    await state.clear()
    await m.answer(AGREEMENT_TEXT, reply_markup={"inline_keyboard": kb})

# ─── ОТПРАВКА PDF ФАЙЛОВ ────────────────────────────────
@router.callback_query(F.data == "policy_pdf")
async def send_policy(c: CallbackQuery):
    await c.message.answer_document(FSInputFile(POLICY_PDF), caption="Политика конфиденциальности")
    await c.answer()

@router.callback_query(F.data == "consent_pdf")
async def send_consent(c: CallbackQuery):
    await c.message.answer_document(FSInputFile(CONSENT_PDF), caption="Текст согласия")
    await c.answer()

@router.callback_query(F.data == "offer_pdf")
async def send_offer(c: CallbackQuery):
    await c.message.answer_document(FSInputFile(OFFER_PDF), caption="Публичная оферта (PDF)")
    await c.answer()

# ─── КНОПКА "СОГЛАСЕН" ─────────────────────────────────
@router.callback_query(F.data == "agree")
async def agree_start(c: CallbackQuery, state: FSMContext):
    await c.answer()
    await state.update_data(consent_status="Согласен")
    await c.message.edit_text("Пожалуйста, введите **ФИО полностью** (одной строкой):")
    await state.set_state(ConsentForm.waiting_fullname)

@router.message(ConsentForm.waiting_fullname)
async def enter_fullname(m: Message, state: FSMContext):
    fullname = m.text.strip()
    if len(fullname.split()) < 2:
        await m.answer("Пожалуйста, укажите ФИО полностью (например: Иванов Иван Иванович).")
        return
    await state.update_data(fullname=fullname)
    await m.answer("Теперь введите **ИНН**:")
    await state.set_state(ConsentForm.waiting_inn)

@router.message(ConsentForm.waiting_inn)
async def enter_inn(m: Message, state: FSMContext):
    inn = m.text.strip()
    if not inn.isdigit() or not (8 <= len(inn) <= 14):
        await m.answer("Проверьте ИНН: должно быть только цифры (8-14 символов). Введите снова:")
        return
    await state.update_data(inn=inn)
    data = await state.get_data()
    await process_final_agree(m, state, data)

async def process_final_agree(m: Message, state: FSMContext, data: dict):
    user = m.from_user
    fullname = data.get("fullname", "")
    inn = data.get("inn", "")
    status = data.get("consent_status", "Согласен")
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Запись в Excel (если файла нет — создать)
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["User ID", "Username", "Имя в Telegram", "ФИО", "ИНН", "Статус", "Время"])
    else:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    ws.append([user.id, user.username, user.full_name, fullname, inn, status, timestamp])
    wb.save(EXCEL_FILE)

    # PDF подтверждение
    pdf_name = f"confirm_{user.id}_{int(datetime.now().timestamp())}.pdf"
    cpdf = canvas.Canvas(pdf_name, pagesize=A4)
    cpdf.setFont("DejaVu", 13)
    cpdf.drawString(90, 800, "Подтверждение согласия")
    cpdf.setFont("DejaVu", 11)
    cpdf.drawString(90, 775, f"User ID: {user.id}")
    cpdf.drawString(90, 760, f"Username: @{user.username or ''}")
    cpdf.drawString(90, 745, f"ФИО: {fullname}")
    cpdf.drawString(90, 730, f"ИНН: {inn}")
    cpdf.drawString(90, 715, f"Статус: {status}")
    cpdf.drawString(90, 700, f"Время: {timestamp}")
    cpdf.drawString(90, 685, f"Актуальные документы: {POLICY_PDF}, {CONSENT_PDF}, {OFFER_PDF}")
    cpdf.save()
    await m.answer_document(FSInputFile(pdf_name), caption="Ваше подтверждение в PDF")
    os.remove(pdf_name)

    await m.answer("Спасибо! Ваш выбор зафиксирован: <b>Согласен</b>\n\n"
                  f"ФИО: <b>{fullname}</b>\nИНН: <b>{inn}</b>", parse_mode="HTML")

    # Уведомление админу (ID, ФИО, ИНН, Telegram)
    try:
        bot = Bot(TOKEN)
        text = (f"✅ Новый ответ!\n"
                f"ID: {user.id}\n"
                f"ФИО: {fullname}\n"
                f"ИНН: {inn}\n"
                f"Имя в Telegram: {user.full_name}\n"
                f"Username: @{user.username or ''}\n"
                f"Время: {timestamp}")
        await bot.send_message(ADMIN_ID, text)
    except Exception as e:
        logging.error(f"Не удалось отправить админу: {e}")

    await state.clear()

# ─── КНОПКА "НЕ СОГЛАСЕН" ───────────────────────────────
@router.callback_query(F.data == "disagree")
async def disagree_handler(c: CallbackQuery):
    user = c.from_user
    status = "Не согласен"
    timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    # Запись в Excel
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.append(["User ID", "Username", "Имя в Telegram", "ФИО", "ИНН", "Статус", "Время"])
    else:
        wb = openpyxl.load_workbook(EXCEL_FILE)
        ws = wb.active
    ws.append([user.id, user.username, user.full_name, "", "", status, timestamp])
    wb.save(EXCEL_FILE)
    await c.message.edit_text("Отказ зафиксирован. Если передумаете — отправьте /start и заполните заново.")
    await c.answer()

# ─── ОТЧЁТ ДЛЯ АДМИНА ───────────────────────────────────
@router.message(Command("report"))
async def report(m: Message):
    if m.from_user.id != ADMIN_ID:
        await m.answer("⛔ У вас нет доступа")
        return
    if not os.path.exists(EXCEL_FILE):
        await m.answer("Файл ещё не создан")
        return
    ts = datetime.now().strftime("%Y%m%d_%H%M%S")
    temp_name = f"consents_{ts}.xlsx"
    shutil.copy(EXCEL_FILE, temp_name)
    await m.answer_document(FSInputFile(temp_name), caption="📊 Отчёт по согласиям")
    os.remove(temp_name)

# ─── ЗАПУСК WEBHOOK ─────────────────────────────────────
from aiohttp import web
async def on_startup(bot: Bot):
    pass
async def on_shutdown(bot: Bot):
    await bot.delete_webhook()

WEBHOOK_HOST = "https://telegram-bot-hdtw.onrender.com"
WEBHOOK_PATH = "/webhook"
WEBHOOK_URL = f"{WEBHOOK_HOST}{WEBHOOK_PATH}"

async def main():
    bot = Bot(TOKEN)
    dp = Dispatcher(storage=storage)
    dp.include_router(router)
    app = web.Application()
    app["bot"] = bot
    app.router.add_post(WEBHOOK_PATH, dp.webhook_handler)
    runner = web.AppRunner(app)
    await runner.setup()
    site = web.TCPSite(runner, "0.0.0.0", int(os.getenv("PORT", 10000)))
    await site.start()
    await bot.set_webhook(WEBHOOK_URL)
    logging.info(f"Webhook запущен: {WEBHOOK_URL}")
    while True:
        await asyncio.sleep(3600)

if __name__ == "__main__":
    import asyncio
    asyncio.run(main())
